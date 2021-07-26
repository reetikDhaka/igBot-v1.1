import openpyxl

def delete_a_data(sheet,upto_row):
    i = 1
    count = 0
    while i <= upto_row:
        #print(f'i = {i}\tcell value(i,1) is {sheet.cell(row=i, column=5).value}')
        if sheet.cell(row=i, column=5).value is True:
            sheet.delete_rows(i, 1)
            print('deleted')
            count += 1
        else:
            i += 1
    print(count)
    wb_obj.save("D:\\follower data\\datarefine.xlsx")
    #wb.save("D:\\follower data\\datatriigred.xlsx")

def show_data(sheet ,start, upto_number):
    for i in range(start, upto_number):
        cell = sheet.cell(i, 2)
        print(f' "{cell.value}", ')




path2 = 'D:\\follower data\\datatriigred.xlsx'
path =  'D:\\follower data\\datarefine.xlsx'
wb_obj = openpyxl.load_workbook(path)
wb = openpyxl.load_workbook(path2)
sheet1 = wb['triggredinsaanfollowerdata']
sheet = wb_obj['Sheet1']
show_data(sheet,600,900)
#delete_a_data(sheet,5000)

